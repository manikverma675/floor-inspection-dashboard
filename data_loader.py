"""
Data loading and shaping for the Floor Inspection dashboard.

Source of truth for Zone -> Bin mapping and per-zone required checks is the
uploaded config workbook ("Zone Config Table-new (1).xlsx").

Inspection records come from the Dataverse CSV exports (fi_*.csv). They are
linked to a zone via:  bin/checklist -> fi_reportid -> master -> zone,
but for bin->zone assignment we ALWAYS use the config workbook.
"""

from __future__ import annotations

import base64
import os
import shutil
import subprocess
from datetime import datetime
from functools import lru_cache

import pandas as pd
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_XLSX_NAME = "Zone Config Table-new (1).xlsx"

# Data (the fi_*.csv exports + the Zone Config workbook) lives in a PRIVATE repo
# and is deliberately not shipped with this public code. Locally the files sit
# next to this module; on Streamlit Cloud we fetch them at runtime using a
# read-only token supplied via the GH_TOKEN secret / env var.
DATA_REPO = os.environ.get("DATA_REPO", "manikverma675/floor-inspection-data")
_CACHE_DIR = os.path.join(HERE, ".data_cache")


def _fetch_private_data(dest: str) -> None:
    """Clone the private data repo into `dest` using GH_TOKEN (read-only)."""
    token = os.environ.get("GH_TOKEN", "").strip()
    if not token:
        raise RuntimeError(
            "Data files are not present locally and no GH_TOKEN was provided. "
            "On Streamlit Cloud, add a GH_TOKEN secret (a fine-grained token with "
            "read-only Contents access to the private data repo)."
        )
    if os.path.exists(dest):
        shutil.rmtree(dest, ignore_errors=True)
    # HTTP Basic auth via an extra header (works for classic PATs, fine-grained
    # PATs and OAuth tokens). Kept out of the URL so it is never written to
    # .git/config, and never surfaced in error text.
    basic = base64.b64encode(f"x-access-token:{token}".encode()).decode()
    cmd = [
        "git", "-c", f"http.extraheader=Authorization: Basic {basic}",
        "clone", "--depth", "1",
        f"https://github.com/{DATA_REPO}.git", dest,
    ]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(
            f"Failed to fetch private data repo '{DATA_REPO}'. Check that the "
            "GH_TOKEN secret is valid and has read access. "
            f"(git exit {exc.returncode})"
        ) from None


@lru_cache(maxsize=1)
def data_dir() -> str:
    """Directory holding the data files: local folder in dev, fetched cache on cloud."""
    if os.path.exists(os.path.join(HERE, CONFIG_XLSX_NAME)):
        return HERE
    if not os.path.exists(os.path.join(_CACHE_DIR, CONFIG_XLSX_NAME)):
        _fetch_private_data(_CACHE_DIR)
    return _CACHE_DIR


def _data_path(name: str) -> str:
    return os.path.join(data_dir(), name)


# Resolved lazily so importing this module never triggers a network fetch.
def _config_xlsx() -> str:
    return _data_path(CONFIG_XLSX_NAME)

# The four per-bin quality checks (Dataverse "*text" columns hold Pass/Fail).
BIN_CHECKS = {
    "fi_aislecleantext": "Aisle Clean",
    "fi_boundarychecktext": "Boundary Check",
    "fi_laneorganisationtext": "Lane Organisation",
    "fi_qrcodestext": "QR Code Readable",
}

# Canonical names for the zone-level safety checks (match the config matrix).
CHECK_BIN = "Bin Quality Inspection"
CHECK_EMERG = "Emergency Exit/Lighting Check"
CHECK_FIRE = "Fire Extinguisher Inspection"
CHECK_EYEWASH = "Eye Wash Station Inspection"
CHECK_DOCK = "Dock Security and Safety Inspection"
CHECK_ELEC = "Electrical Safety"
CHECK_MOISTURE = "Moisture Control"
CHECK_5S = "5S Board"

SAFETY_CHECKS = [
    CHECK_EMERG, CHECK_FIRE, CHECK_EYEWASH, CHECK_DOCK,
    CHECK_ELEC, CHECK_MOISTURE, CHECK_5S,
]


def _read_csv(name: str) -> pd.DataFrame:
    return pd.read_csv(_data_path(name), encoding="utf-8-sig", dtype=str).fillna("")


def report_id_to_dt(rid: str):
    """QA + YYYYMMDDHHMMSS -> datetime (None if not parseable)."""
    try:
        return datetime.strptime(str(rid)[2:], "%Y%m%d%H%M%S")
    except Exception:
        return None


# --------------------------------------------------------------------------- #
# Config workbook
# --------------------------------------------------------------------------- #
@lru_cache(maxsize=1)
def load_config():
    """Return (bin_to_zone: dict, zone_to_bins: dict, required: DataFrame)."""
    wb = openpyxl.load_workbook(_config_xlsx(), data_only=True)

    # ---- Sheet 2: Zone -> Bin (mixed layout; header + first bin may share a row)
    ws = wb["Sheet2"]
    bin_to_zone: dict[str, str] = {}
    zone_to_bins: dict[str, list[str]] = {}
    cur = None
    for zc, bc in ws.iter_rows(values_only=True):
        zv = str(zc).strip() if zc is not None else ""
        bv = str(bc).strip() if bc is not None else ""
        if zv.lower().startswith("zone") and zv.lower() != "zone":
            cur = zv
            zone_to_bins.setdefault(cur, [])
        if bv and bv.lower() != "bin" and cur:
            zone_to_bins[cur].append(bv)
            bin_to_zone[bv] = cur

    # ---- Sheet 1: required checks per zone (Y/N matrix)
    ws1 = wb["new config table "]
    rows = list(ws1.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        zone = str(r[0]).strip()
        # normalise "Zone1" -> "Zone 1"
        if zone.lower().startswith("zone") and " " not in zone:
            zone = "Zone " + zone[4:]
        rec = {"Zone": zone}
        for i in range(1, len(header)):
            rec[header[i]] = str(r[i]).strip().upper() == "Y" if r[i] is not None else False
        records.append(rec)
    required = pd.DataFrame(records)

    return bin_to_zone, zone_to_bins, required


def zone_sort_key(z: str) -> int:
    try:
        return int(str(z).split()[-1])
    except Exception:
        return 999


# --------------------------------------------------------------------------- #
# Master report table (zone visits)
# --------------------------------------------------------------------------- #
@lru_cache(maxsize=1)
def load_reports() -> pd.DataFrame:
    """One row per inspection report (zone visit) with derived date + zone."""
    m = _read_csv("fi_floorinspectionmasters.csv")
    m = m[m["fi_reportid"] != ""].copy()
    m["zone"] = m["fi_zoneconfigtext"].str.strip()
    m["date"] = m["fi_reportid"].map(report_id_to_dt)
    m["result"] = m["fi_inspectionresulttext"].str.strip()
    m["issues"] = m["fi_observedissues"].str.strip()
    m["comments"] = m["fi_additionalcomments"].str.strip()
    return m[["fi_reportid", "zone", "date", "result", "issues", "comments"]]


def _report_zone_map() -> dict[str, str]:
    """reportid -> zone from the master table (used only to label safety rows)."""
    reps = load_reports()
    return dict(zip(reps["fi_reportid"], reps["zone"]))


# --------------------------------------------------------------------------- #
# Bin inspection records (the rich table)
# --------------------------------------------------------------------------- #
@lru_cache(maxsize=1)
def load_bin_inspections() -> pd.DataFrame:
    """
    One row per bin inspection occurrence. Zone comes from the CONFIG file
    (bin -> zone), falling back to '(unmapped)' if a bin isn't in the config.
    """
    df = _read_csv("fi_binqualityinspections.csv")
    df = df[df["fi_binlocationtext"] != ""].copy()
    bin_to_zone, _, _ = load_config()

    df["bin"] = df["fi_binlocationtext"].str.strip()
    df["zone"] = df["bin"].map(bin_to_zone).fillna("(unmapped)")
    df["date"] = df["fi_reportid"].map(report_id_to_dt)

    # normalise the four checks to Pass/Fail/(blank)
    for col, label in BIN_CHECKS.items():
        df[label] = df[col].str.strip()

    check_labels = list(BIN_CHECKS.values())
    df["n_fail"] = (df[check_labels] == "Fail").sum(axis=1)
    df["passed_all"] = df["n_fail"] == 0
    df["failed_checks"] = df[check_labels].apply(
        lambda r: [lbl for lbl in check_labels if r[lbl] == "Fail"], axis=1
    )
    return df


def failed_bins_long() -> pd.DataFrame:
    """One row per (bin inspection, failed check) — for charts/ledgers."""
    df = load_bin_inspections()
    rows = []
    for _, r in df.iterrows():
        for chk in r["failed_checks"]:
            rows.append({
                "date": r["date"], "zone": r["zone"], "bin": r["bin"],
                "check": chk, "fi_reportid": r["fi_reportid"],
            })
    return pd.DataFrame(rows, columns=["date", "zone", "bin", "check", "fi_reportid"])


# --------------------------------------------------------------------------- #
# Zone-level safety checks -> tidy status per (report, check)
# --------------------------------------------------------------------------- #
def _status(passed: bool | None, recorded: bool) -> str:
    if not recorded:
        return "Not recorded"
    if passed is None:
        return "Recorded"
    return "Pass" if passed else "Fail"


@lru_cache(maxsize=1)
def load_safety_status() -> pd.DataFrame:
    """
    Tidy frame: fi_reportid, zone, check, status in
    {Pass, Fail, Recorded, Not recorded}.
    Result is read from the '*text' fields (Pass/Fail); the boolean 'fi_passed'
    columns are unreliable and deliberately ignored.
    """
    zmap = _report_zone_map()
    out = []

    def add(rid, check, passed, recorded):
        out.append({
            "fi_reportid": rid, "zone": zmap.get(rid, "(no master)"),
            "check": check, "status": _status(passed, recorded),
        })

    # Electrical — real Pass/Fail signal lives in the text field
    for _, r in _read_csv("fi_electricalsafeties.csv").iterrows():
        t = r["fi_electricalsafetytext"].strip()
        add(r["fi_reportid"], CHECK_ELEC,
            (t == "Pass") if t in ("Pass", "Fail") else None, bool(t))

    # Eyewash / Moisture / 5S — text field carries Pass
    for fname, col, check in [
        ("fi_eyewashstationinspections.csv", "fi_eyewashstationinspectiontext", CHECK_EYEWASH),
        ("fi_moisturecontrols.csv", "fi_moisturecontroltext", CHECK_MOISTURE),
        ("fi_sboards.csv", "fi_sboardtext", CHECK_5S),
    ]:
        for _, r in _read_csv(fname).iterrows():
            t = r[col].strip()
            add(r["fi_reportid"], check,
                (t == "Pass") if t in ("Pass", "Fail") else None, bool(t))

    # Fire extinguisher — "needs attention" count; pass when recorded & 0 needing attention
    for _, r in _read_csv("fi_fireextinguisherinspections.csv").iterrows():
        good = r["fi_conditiongoodtext"].strip()
        needs = r["fi_conditionneedsattentiontext"].strip()
        recorded = bool(good) or bool(needs)
        passed = None
        if recorded:
            passed = (needs in ("", "0"))
        add(r["fi_reportid"], CHECK_FIRE, passed, recorded)

    # Emergency exit lighting — recorded if any brightness/condition captured
    for _, r in _read_csv("fi_emergencyexitlightingchecks.csv").iterrows():
        recorded = any(r[c].strip() for c in [
            "fi_areafreefromanyobstructions", "fi_brightnesslevelbrighttext",
            "fi_conditionworkingtext"])
        needs_repl = r["fi_conditionneedsreplacementstext"].strip()
        passed = None
        if recorded:
            passed = (needs_repl in ("", "0"))
        add(r["fi_reportid"], CHECK_EMERG, passed, recorded)

    # Dock security — three sub-checks coded 122430000 (=Pass) when present
    for _, r in _read_csv("fi_docksecurityandsafetyinspections.csv").iterrows():
        recorded = any(r[c].strip() for c in
                       ["fi_chockblocks", "fi_droptrailers", "fi_unuseddockdoors"])
        add(r["fi_reportid"], CHECK_DOCK, True if recorded else None, recorded)

    return pd.DataFrame(out)


# --------------------------------------------------------------------------- #
# Completeness: required checks (config) vs what was recorded
# --------------------------------------------------------------------------- #
def completeness_by_report() -> pd.DataFrame:
    """
    Per report: how many of the zone's REQUIRED checks were actually recorded.
    Bin Quality counts as recorded if the report has any bin rows.
    """
    reps = load_reports()
    _, _, required = load_config()
    req = required.set_index("Zone")
    safety = load_safety_status()
    bins = load_bin_inspections()
    bin_reports = set(bins["fi_reportid"])

    recorded_map = {
        (row["fi_reportid"], row["check"]): row["status"] != "Not recorded"
        for _, row in safety.iterrows()
    }

    rows = []
    for _, rep in reps.iterrows():
        zone = rep["zone"]
        if zone not in req.index:
            continue
        req_row = req.loc[zone]
        required_checks = [c for c in req_row.index if bool(req_row[c])]
        done = 0
        missing = []
        for chk in required_checks:
            if chk == CHECK_BIN:
                ok = rep["fi_reportid"] in bin_reports
            else:
                ok = recorded_map.get((rep["fi_reportid"], chk), False)
            if ok:
                done += 1
            else:
                missing.append(chk)
        rows.append({
            "fi_reportid": rep["fi_reportid"], "zone": zone, "date": rep["date"],
            "required": len(required_checks), "completed": done,
            "missing_checks": missing,
            "pct": (done / len(required_checks) * 100) if required_checks else 100.0,
        })
    return pd.DataFrame(rows)


if __name__ == "__main__":
    # quick smoke test
    b2z, z2b, req = load_config()
    print("bins mapped:", len(b2z), "zones:", len(z2b))
    print("required matrix:\n", req)
    bi = load_bin_inspections()
    print("\nbin inspections:", len(bi), "| unmapped:",
          (bi["zone"] == "(unmapped)").sum())
    print("failed-check rows:", len(failed_bins_long()))
    print("\nsafety status counts:\n", load_safety_status()["status"].value_counts())
    print("\ncompleteness:\n", completeness_by_report()[["zone", "date", "required", "completed", "pct"]])
